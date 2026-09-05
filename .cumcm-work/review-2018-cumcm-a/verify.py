#!/usr/bin/env python3
"""Recompute every numeric claim in review.json from the official 2018 A attachment.

Read-only. Run:
    python .cumcm-work/review-2018-cumcm-a/verify.py
"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

ATTACH = Path(r"F:\CUMCM\.cumcm-work\blind-source-2018-a\CUMCM.xlsx")
LAYERS = ("I", "II", "III", "IV")
PAPER_DIFFUSIVITY = {"I": 1.98e-7, "II": 2.04e-7, "III": 3.51e-7, "IV": 2.6311e-7}
CODE_COEFFICIENT = {"I": 0.082, "II": 0.37, "III": 0.045}


def read_attachment() -> tuple[dict[str, dict[str, float]], list[tuple[float, float]]]:
    book = load_workbook(ATTACH, read_only=True, data_only=True)
    properties: dict[str, dict[str, float]] = {}
    for row in book["附件1"].iter_rows(values_only=True):
        label = str(row[0] or "").strip()
        if label[:-1] in LAYERS and label.endswith("层"):
            properties[label[:-1]] = {
                "rho": float(row[1]), "c": float(row[2]), "lambda": float(row[3]),
                "thickness_raw": str(row[4]), "attachment_diffusivity": float(row[5]),
            }
    series: list[tuple[float, float]] = []
    for index, row in enumerate(book["附件2"].iter_rows(values_only=True)):
        if index < 2:
            continue
        try:
            series.append((float(row[0]), float(row[1])))
        except (TypeError, ValueError):
            continue
    book.close()
    return properties, series


def steady_state(env: float, skin: float, d2_mm: float, d4_mm: float,
                 properties: dict[str, dict[str, float]]) -> tuple[float, list[float], float]:
    """Series slabs, no generation: fixing both end temperatures fixes only the flux."""
    thickness = {"I": 0.6e-3, "II": d2_mm * 1e-3, "III": 3.6e-3, "IV": d4_mm * 1e-3}
    resistance = sum(thickness[k] / properties[k]["lambda"] for k in LAYERS)
    flux = (env - skin) / resistance
    interfaces = [env]
    for key in LAYERS:
        interfaces.append(interfaces[-1] - flux * thickness[key] / properties[key]["lambda"])
    return flux, interfaces, resistance


def main() -> int:
    properties, series = read_attachment()
    report: dict[str, object] = {"attachment": str(ATTACH), "samples": len(series)}

    # F1 / F3 -- diffusivity
    diffusivity = []
    for key in LAYERS:
        p = properties[key]
        computed = p["lambda"] / (p["rho"] * p["c"])
        diffusivity.append({
            "layer": key, "computed": computed,
            "attachment": p["attachment_diffusivity"],
            "paper": PAPER_DIFFUSIVITY[key],
            "attachment_matches_computed": abs(computed - p["attachment_diffusivity"]) / computed < 1e-6,
            "paper_relative_error": abs(computed - PAPER_DIFFUSIVITY[key]) / computed,
            "code_coefficient_ratio": (CODE_COEFFICIENT[key] / computed) if key in CODE_COEFFICIENT else None,
        })
    report["F1_F3_diffusivity"] = diffusivity

    # F4 -- transient facts straight from the measurement
    times = [t for t, _ in series]
    values = [u for _, u in series]
    quasi_steady = next(
        (times[i] for i in range(60, len(values)) if abs(values[i] - values[i - 60]) < 0.01), None)
    thresholds = {}
    for threshold in (44.0, 47.0):
        above = [t for t, u in series if u > threshold]
        thresholds[str(threshold)] = {
            "first_s": above[0] if above else None,
            "duration_s": len(above),
        }
    report["F4_F8_measurement"] = {
        "start_c": values[0], "final_c": values[-1], "max_c": max(values),
        "quasi_steady_s": quasi_steady,
        "last_10min_spread_c": max(values[-600:]) - min(values[-600:]),
        "above": thresholds,
    }

    # F2 -- the steady-state system leaves d2 free
    scan = []
    for d2 in (0.6, 2.0, 5.5, 10.0, 25.0):
        flux, interfaces, resistance = steady_state(65.0, 47.0, d2, 5.5, properties)
        scan.append({"d2_mm": d2, "resistance": resistance, "flux_w_m2": flux,
                     "interfaces_c": [round(v, 4) for v in interfaces]})
    report["F2_steady_state_scan"] = {
        "note": "每个 d2 都有自洽解；4 个方程对 theta2,theta3,theta4,q,d2 共 5 个未知量。",
        "rows": scan,
    }

    # F5 -- reported thicknesses against the attachment bounds
    report["F5_bounds"] = {
        "II_raw": properties["II"]["thickness_raw"], "IV_raw": properties["IV"]["thickness_raw"],
        "abstract_d4_mm": 5.4, "body_d4_mm": 6.4,
        "body_d4_equals_upper_bound": properties["IV"]["thickness_raw"].endswith("6.4"),
    }

    print(json.dumps(report, ensure_ascii=False, indent=1))
    failures = [d["layer"] for d in diffusivity if d["paper_relative_error"] > 0.05]
    print(f"\ndiffusivity mismatches vs attachment: {failures or 'none'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
