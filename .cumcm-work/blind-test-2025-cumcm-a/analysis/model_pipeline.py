from __future__ import annotations

import json
import math
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from scipy.optimize import differential_evolution


SEED = 20250814
G = 9.8
CLOUD_RADIUS = 10.0
CLOUD_LIFE = 20.0
CLOUD_SINK = 3.0

ROOT = Path(__file__).resolve().parents[1]
RESULT_DIR = ROOT / "analysis" / "results"
FIGURE_DIR = ROOT / "paper" / "figures"
DATA_DIR = ROOT / "data"

UAVS = {
    "FY1": np.array([17800.0, 0.0, 1800.0]),
    "FY2": np.array([12000.0, 1400.0, 1400.0]),
    "FY3": np.array([6000.0, -3000.0, 700.0]),
    "FY4": np.array([11000.0, 2000.0, 1800.0]),
    "FY5": np.array([13000.0, -2000.0, 1300.0]),
}

MISSILES = {
    "M1": np.array([20000.0, 0.0, 2000.0]),
    "M2": np.array([19000.0, 600.0, 2100.0]),
    "M3": np.array([18000.0, -600.0, 1900.0]),
}


def target_points() -> np.ndarray:
    points = [np.array([0.0, 200.0, 5.0])]
    for z in (0.0, 10.0):
        for angle in np.linspace(0.0, 2.0 * np.pi, 8, endpoint=False):
            points.append(
                np.array([7.0 * np.cos(angle), 200.0 + 7.0 * np.sin(angle), z])
            )
    return np.asarray(points)


TARGET_POINTS = target_points()


@dataclass(frozen=True)
class Action:
    uav: str
    missile: str
    heading: float
    speed: float
    release_time: float
    delay: float

    @property
    def detonation_time(self) -> float:
        return self.release_time + self.delay


def missile_velocity(missile: str) -> np.ndarray:
    initial = MISSILES[missile]
    return -300.0 * initial / np.linalg.norm(initial)


def impact_time(missile: str) -> float:
    return float(np.linalg.norm(MISSILES[missile]) / 300.0)


def time_grid(missile: str, dt: float) -> np.ndarray:
    return np.arange(0.0, impact_time(missile) + dt / 2.0, dt)


def action_points(action: Action) -> tuple[np.ndarray, np.ndarray]:
    start = UAVS[action.uav]
    horizontal = np.array(
        [
            action.speed * np.cos(action.heading),
            action.speed * np.sin(action.heading),
            0.0,
        ]
    )
    release = start + horizontal * action.release_time
    detonation = start + horizontal * action.detonation_time
    detonation = detonation + np.array([0.0, 0.0, -0.5 * G * action.delay**2])
    return release, detonation


def distance_profile(
    action: Action, grid: np.ndarray
) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
    """Return target-wide maximum ray distance, valid-time mask and occlusion mask."""
    release, detonation = action_points(action)
    del release
    valid = (
        (grid >= action.detonation_time)
        & (grid <= action.detonation_time + CLOUD_LIFE)
        & (grid <= impact_time(action.missile))
        & (detonation[2] >= 0.0)
    )
    max_distance = np.full(grid.shape, np.inf, dtype=float)
    occluded = np.zeros(grid.shape, dtype=bool)
    if not np.any(valid):
        return max_distance, valid, occluded

    times = grid[valid]
    missile = MISSILES[action.missile][None, :] + times[:, None] * missile_velocity(
        action.missile
    )
    cloud = detonation[None, :] + np.column_stack(
        (
            np.zeros(times.size),
            np.zeros(times.size),
            -CLOUD_SINK * (times - action.detonation_time),
        )
    )

    ray = TARGET_POINTS[None, :, :] - missile[:, None, :]
    cloud_ray = cloud[:, None, :] - missile[:, None, :]
    fraction = np.sum(cloud_ray * ray, axis=2) / np.sum(ray * ray, axis=2)
    nearest = missile[:, None, :] + np.clip(fraction, 0.0, 1.0)[:, :, None] * ray
    distance = np.linalg.norm(cloud[:, None, :] - nearest, axis=2)
    local_max = np.max(distance, axis=1)
    local_ok = (
        (local_max <= CLOUD_RADIUS)
        & np.all(fraction >= 0.0, axis=1)
        & np.all(fraction <= 1.0, axis=1)
    )
    max_distance[valid] = local_max
    occluded[valid] = local_ok
    return max_distance, valid, occluded


def mask_intervals(grid: np.ndarray, mask: np.ndarray) -> list[list[float]]:
    indices = np.flatnonzero(mask)
    if indices.size == 0:
        return []
    intervals: list[list[float]] = []
    start = previous = int(indices[0])
    for index in indices[1:]:
        index = int(index)
        if index > previous + 1:
            intervals.append([float(grid[start]), float(grid[previous])])
            start = index
        previous = index
    intervals.append([float(grid[start]), float(grid[previous])])
    return intervals


def mask_duration(mask: np.ndarray, dt: float) -> float:
    return float(np.count_nonzero(mask) * dt)


def evaluate_action(action: Action, dt: float = 0.01) -> dict:
    grid = time_grid(action.missile, dt)
    max_distance, valid, mask = distance_profile(action, grid)
    release, detonation = action_points(action)
    finite = max_distance[valid]
    return {
        "uav": action.uav,
        "missile": action.missile,
        "heading_deg": float(np.degrees(action.heading) % 360.0),
        "speed": float(action.speed),
        "release_time": float(action.release_time),
        "delay": float(action.delay),
        "detonation_time": float(action.detonation_time),
        "release_point": release.tolist(),
        "detonation_point": detonation.tolist(),
        "effective_duration": mask_duration(mask, dt),
        "intervals": mask_intervals(grid, mask),
        "minimum_worst_ray_distance": float(np.min(finite)) if finite.size else None,
    }


def action_from_vector(uav: str, missile: str, vector: Iterable[float]) -> Action:
    heading, speed, release_time, delay = (float(value) for value in vector)
    return Action(uav, missile, heading, speed, release_time, delay)


def optimize_single(uav: str, missile: str, seed_offset: int = 0) -> Action:
    grid = time_grid(missile, 0.05)
    max_delay = min(18.0, math.sqrt(2.0 * UAVS[uav][2] / G) - 0.05)
    release_upper = min(35.0, impact_time(missile) - 0.5)

    def objective(vector: np.ndarray) -> float:
        action = action_from_vector(uav, missile, vector)
        max_distance, valid, mask = distance_profile(action, grid)
        if not np.any(valid):
            return 1.0e6
        duration = mask_duration(mask, 0.05)
        closest = float(np.min(max_distance[valid]))
        # Distance guides the search into the narrow feasible tube; duration decides
        # among candidates once the whole target is hidden.
        return closest - 100.0 * duration

    result = differential_evolution(
        objective,
        bounds=[
            (0.0, 2.0 * np.pi),
            (70.0, 140.0),
            (0.0, release_upper),
            (0.2, max_delay),
        ],
        seed=SEED + seed_offset,
        popsize=12,
        maxiter=100,
        tol=0.002,
        polish=False,
        workers=1,
        updating="immediate",
    )
    return action_from_vector(uav, missile, result.x)


def candidate_actions_for_path(
    uav: str,
    missile: str,
    heading: float,
    speed: float,
    release_step: float,
    delay_step: float,
) -> list[tuple[Action, np.ndarray]]:
    grid = time_grid(missile, 0.05)
    max_delay = min(18.0, math.sqrt(2.0 * UAVS[uav][2] / G) - 0.05)
    release_upper = min(35.0, impact_time(missile) - 0.5)
    candidates: list[tuple[Action, np.ndarray]] = []
    for release_time in np.arange(0.0, release_upper + release_step / 2, release_step):
        for delay in np.arange(0.25, max_delay + delay_step / 2, delay_step):
            action = Action(uav, missile, heading, speed, float(release_time), float(delay))
            _, _, mask = distance_profile(action, grid)
            if np.any(mask):
                candidates.append((action, mask))
    candidates.sort(key=lambda item: int(np.count_nonzero(item[1])), reverse=True)
    return candidates


def select_path_plan(
    candidates: list[tuple[Action, np.ndarray]], missile: str, max_bombs: int = 3
) -> tuple[list[Action], np.ndarray]:
    grid = time_grid(missile, 0.05)
    union = np.zeros(grid.shape, dtype=bool)
    selected: list[Action] = []
    for _ in range(max_bombs):
        best: tuple[tuple[int, int, float], Action, np.ndarray] | None = None
        for action, mask in candidates:
            if any(abs(action.release_time - old.release_time) < 1.0 - 1.0e-9 for old in selected):
                continue
            gain = int(np.count_nonzero(mask & ~union))
            overlap = int(np.count_nonzero(mask & union))
            score = (gain, -overlap, -action.release_time)
            if gain > 0 and (best is None or score > best[0]):
                best = (score, action, mask)
        if best is None:
            break
        _, action, mask = best
        selected.append(action)
        union |= mask
    selected.sort(key=lambda action: action.release_time)
    return selected, union


def refine_path_plan(
    uav: str, missile: str, heading: float, speed: float
) -> tuple[list[Action], float]:
    coarse_candidates = candidate_actions_for_path(
        uav, missile, heading, speed, release_step=0.5, delay_step=0.5
    )
    coarse_plan, _ = select_path_plan(coarse_candidates, missile)
    if not coarse_plan:
        return [], 0.0

    release_values: set[float] = set()
    delay_values: set[float] = set()
    for action in coarse_plan:
        for value in np.arange(max(0.0, action.release_time - 0.75), action.release_time + 0.76, 0.25):
            release_values.add(round(float(value), 6))
        for value in np.arange(max(0.25, action.delay - 0.75), action.delay + 0.76, 0.25):
            delay_values.add(round(float(value), 6))

    grid = time_grid(missile, 0.05)
    max_delay = min(18.0, math.sqrt(2.0 * UAVS[uav][2] / G) - 0.05)
    refined: list[tuple[Action, np.ndarray]] = []
    for release_time in sorted(release_values):
        for delay in sorted(delay_values):
            if delay > max_delay:
                continue
            action = Action(uav, missile, heading, speed, release_time, delay)
            _, _, mask = distance_profile(action, grid)
            if np.any(mask):
                refined.append((action, mask))
    plan, union = select_path_plan(refined, missile)
    return plan, mask_duration(union, 0.05)


def union_evaluation(actions: list[Action], missile: str, dt: float = 0.01) -> dict:
    grid = time_grid(missile, dt)
    union = np.zeros(grid.shape, dtype=bool)
    individual = []
    for action in actions:
        _, _, mask = distance_profile(action, grid)
        union |= mask
        individual.append(evaluate_action(action, dt))
    return {
        "missile": missile,
        "union_duration": mask_duration(union, dt),
        "union_intervals": mask_intervals(grid, union),
        "actions": individual,
    }


def write_result1(plan: list[Action]) -> None:
    workbook = load_workbook(DATA_DIR / "result1-template.xlsx")
    sheet = workbook.active
    for row, action in zip(range(2, 5), plan):
        record = evaluate_action(action)
        release = record["release_point"]
        detonation = record["detonation_point"]
        values = [
            record["heading_deg"],
            record["speed"],
            row - 1,
            *release,
            *detonation,
            record["effective_duration"],
        ]
        for column, value in enumerate(values, 1):
            sheet.cell(row=row, column=column, value=float(value))
    workbook.save(RESULT_DIR / "result1.xlsx")


def write_result2(actions: list[Action]) -> None:
    workbook = load_workbook(DATA_DIR / "result2-template.xlsx")
    sheet = workbook.active
    by_uav = {action.uav: action for action in actions}
    for row, uav in zip(range(2, 5), ("FY1", "FY2", "FY3")):
        action = by_uav[uav]
        record = evaluate_action(action)
        values = [
            uav,
            record["heading_deg"],
            record["speed"],
            *record["release_point"],
            *record["detonation_point"],
            record["effective_duration"],
        ]
        for column, value in enumerate(values, 1):
            sheet.cell(row=row, column=column, value=value)
    workbook.save(RESULT_DIR / "result2.xlsx")


def write_result3(assignment: dict[str, tuple[str, list[Action]]]) -> None:
    workbook = load_workbook(DATA_DIR / "result3-template.xlsx")
    sheet = workbook.active
    row = 2
    for uav in UAVS:
        missile, actions = assignment[uav]
        for bomb_number in range(1, 4):
            if bomb_number <= len(actions):
                record = evaluate_action(actions[bomb_number - 1])
                values = [
                    uav,
                    record["heading_deg"],
                    record["speed"],
                    bomb_number,
                    *record["release_point"],
                    *record["detonation_point"],
                    record["effective_duration"],
                    missile,
                ]
                for column, value in enumerate(values, 1):
                    sheet.cell(row=row, column=column, value=value)
            row += 1
    workbook.save(RESULT_DIR / "result3.xlsx")


def plot_distance(action: Action, filename: str, title: str) -> None:
    grid = time_grid(action.missile, 0.01)
    max_distance, valid, mask = distance_profile(action, grid)
    figure, axis = plt.subplots(figsize=(8.2, 4.2))
    axis.plot(grid[valid], max_distance[valid], color="#1f5a78", linewidth=1.6)
    axis.axhline(CLOUD_RADIUS, color="#b23a32", linestyle="--", label="10 m threshold")
    if np.any(mask):
        axis.fill_between(
            grid,
            0,
            CLOUD_RADIUS,
            where=mask,
            color="#4f8f5b",
            alpha=0.2,
            label="effective interval",
        )
    axis.set_xlabel("time / s")
    axis.set_ylabel("maximum ray distance / m")
    axis.set_title(title)
    axis.grid(alpha=0.25)
    axis.legend()
    figure.tight_layout()
    figure.savefig(FIGURE_DIR / filename)
    plt.close(figure)


def plot_plan_timeline(plans: dict[str, tuple[str, list[Action]]], filename: str) -> None:
    figure, axis = plt.subplots(figsize=(9.2, 5.0))
    colors = {"M1": "#315f76", "M2": "#ad6a32", "M3": "#577b55"}
    y = 0
    ticks = []
    labels = []
    for uav, (missile, actions) in plans.items():
        for action in actions:
            record = evaluate_action(action)
            for start, end in record["intervals"]:
                axis.barh(y, end - start, left=start, height=0.55, color=colors[missile])
        ticks.append(y)
        labels.append(f"{uav}->{missile}")
        y += 1
    axis.set_yticks(ticks, labels)
    axis.set_xlabel("time after task assignment / s")
    axis.set_title("Effective shielding intervals of the selected actions")
    axis.grid(axis="x", alpha=0.25)
    figure.tight_layout()
    figure.savefig(FIGURE_DIR / filename)
    plt.close(figure)


def main() -> None:
    RESULT_DIR.mkdir(parents=True, exist_ok=True)
    FIGURE_DIR.mkdir(parents=True, exist_ok=True)
    plt.rcParams.update({"font.size": 10, "axes.unicode_minus": False})

    q1 = Action("FY1", "M1", np.pi, 120.0, 1.5, 3.6)
    q1_result = evaluate_action(q1, dt=0.001)
    plot_distance(q1, "q1_distance.pdf", "Problem 1: worst target-ray distance")

    optimized: dict[tuple[str, str], Action] = {}
    optimized[("FY1", "M1")] = optimize_single("FY1", "M1", 1)
    q2 = optimized[("FY1", "M1")]
    q2_result = evaluate_action(q2, dt=0.001)
    plot_distance(q2, "q2_distance.pdf", "Problem 2: optimized worst target-ray distance")

    # A single-bomb path is only a starting point. Nearby headings and speeds are
    # compared by the union created by three release events on the same path.
    path_options: list[tuple[float, float, list[Action]]] = []
    base_heading = q2.heading
    for heading_delta in np.radians(np.arange(-8.0, 8.1, 2.0)):
        for speed in np.arange(70.0, 121.0, 10.0):
            plan, duration = refine_path_plan(
                "FY1", "M1", float(base_heading + heading_delta), float(speed)
            )
            if plan:
                path_options.append((duration, float(speed), plan))
    if not path_options:
        raise RuntimeError("No feasible three-bomb path found for Problem 3")
    _, _, q3_plan = max(path_options, key=lambda item: item[0])
    q3_result = union_evaluation(q3_plan, "M1", dt=0.01)
    write_result1(q3_plan)

    # The remaining single-action searches are reused in Problems 4 and 5.
    seed_offset = 20
    for uav in UAVS:
        for missile in MISSILES:
            if (uav, missile) not in optimized:
                optimized[(uav, missile)] = optimize_single(uav, missile, seed_offset)
                seed_offset += 1

    q4_actions = [optimized[(uav, "M1")] for uav in ("FY1", "FY2", "FY3")]
    q4_result = union_evaluation(q4_actions, "M1", dt=0.01)
    write_result2(q4_actions)

    pair_plans: dict[tuple[str, str], tuple[list[Action], float]] = {}
    for pair, action in optimized.items():
        pair_plans[pair] = refine_path_plan(
            pair[0], pair[1], action.heading, action.speed
        )

    # Each UAV keeps one heading and speed. Enumerating 3^5 assignments is exact
    # over the candidate path plans and preserves the per-UAV three-bomb limit.
    best_score = -1.0
    best_assignment: dict[str, tuple[str, list[Action]]] | None = None
    missile_names = tuple(MISSILES)
    for choices in np.ndindex(*(3 for _ in UAVS)):
        assignment: dict[str, tuple[str, list[Action]]] = {}
        unions = {
            missile: np.zeros(time_grid(missile, 0.05).shape, dtype=bool)
            for missile in missile_names
        }
        for uav, choice in zip(UAVS, choices):
            missile = missile_names[choice]
            plan, _ = pair_plans[(uav, missile)]
            assignment[uav] = (missile, plan)
            grid = time_grid(missile, 0.05)
            for action in plan:
                _, _, mask = distance_profile(action, grid)
                unions[missile] |= mask
        durations = {
            missile: mask_duration(unions[missile], 0.05) for missile in missile_names
        }
        if any(value <= 0.0 for value in durations.values()):
            continue
        score = sum(durations.values())
        if score > best_score:
            best_score = score
            best_assignment = assignment
    if best_assignment is None:
        raise RuntimeError("No assignment covers all three missiles")

    q5_by_missile = {
        missile: union_evaluation(
            [
                action
                for assigned_missile, actions in best_assignment.values()
                if assigned_missile == missile
                for action in actions
            ],
            missile,
            dt=0.01,
        )
        for missile in MISSILES
    }
    write_result3(best_assignment)
    plot_plan_timeline(best_assignment, "q5_timeline.pdf")

    summary = {
        "seed": SEED,
        "criterion": {
            "cloud_radius_m": CLOUD_RADIUS,
            "cloud_life_s": CLOUD_LIFE,
            "cloud_sink_m_per_s": CLOUD_SINK,
            "target_sample_points": int(TARGET_POINTS.shape[0]),
            "target_geometry": "cylinder center plus eight rim points at z=0 and z=10",
            "effective_rule": "all sampled target rays intersect the cloud sphere",
        },
        "problem1": q1_result,
        "problem2": q2_result,
        "problem3": q3_result,
        "problem4": q4_result,
        "problem5": {
            "objective_sum_duration": float(
                sum(item["union_duration"] for item in q5_by_missile.values())
            ),
            "by_missile": q5_by_missile,
            "assignment": {
                uav: {"missile": missile, "bomb_count": len(actions)}
                for uav, (missile, actions) in best_assignment.items()
            },
        },
        "search_boundary": {
            "single_action": "differential evolution with distance guidance and fixed seed",
            "same_path_plan": "0.5 s coarse event grid followed by 0.25 s local refinement",
            "multi_missile_assignment": "exact enumeration over 3^5 candidate-path assignments",
        },
    }
    (RESULT_DIR / "summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
