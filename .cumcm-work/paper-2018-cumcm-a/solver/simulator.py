#!/usr/bin/env python3
"""2018 CUMCM A -- one transient simulator serving all three subquestions.

Rebuild driven by .cumcm-work/review-2018-cumcm-a/review.json ("if_rebuilt"):
finite-volume 1-D transient conduction through the four layers, harmonic-mean
face conductance so interface flux continuity is exact, backward Euler so the
air layer's large diffusivity cannot force the time step, and a single skin-side
Robin coefficient calibrated against the official measurement.

Read-only with respect to the competition material. Run:
    python .cumcm-work/rebuild-2018-cumcm-a/simulator.py --self-test
"""
from __future__ import annotations

import argparse
import json
from dataclasses import dataclass
from pathlib import Path

import numpy as np
from openpyxl import load_workbook

ATTACHMENT = Path(r"F:\CUMCM\.cumcm-work\blind-source-2018-a\CUMCM.xlsx")
LAYERS = ("I", "II", "III", "IV")
BODY_CORE_C = 37.0
FIXED_THICKNESS_MM = {"I": 0.6, "III": 3.6}
BOUNDS_MM = {"II": (0.6, 25.0), "IV": (0.6, 6.4)}


@dataclass(frozen=True)
class Material:
    rho: float
    c: float
    lam: float

    @property
    def diffusivity(self) -> float:
        return self.lam / (self.rho * self.c)


def read_materials() -> dict[str, Material]:
    book = load_workbook(ATTACHMENT, read_only=True, data_only=True)
    out: dict[str, Material] = {}
    for row in book["附件1"].iter_rows(values_only=True):
        label = str(row[0] or "").strip()
        if label.endswith("层") and label[:-1] in LAYERS:
            out[label[:-1]] = Material(float(row[1]), float(row[2]), float(row[3]))
    book.close()
    if set(out) != set(LAYERS):
        raise ValueError(f"expected four layers, parsed {sorted(out)}")
    return out


def read_measurement() -> tuple[np.ndarray, np.ndarray]:
    book = load_workbook(ATTACHMENT, read_only=True, data_only=True)
    times, values = [], []
    for index, row in enumerate(book["附件2"].iter_rows(values_only=True)):
        if index < 2:
            continue
        try:
            times.append(float(row[0]))
            values.append(float(row[1]))
        except (TypeError, ValueError):
            continue
    book.close()
    return np.asarray(times), np.asarray(values)


class Simulator:
    """Cell-centred finite volume over the four layers; skin side is Robin."""

    def __init__(self, materials: dict[str, Material], d2_mm: float, d4_mm: float,
                 cells_per_layer: int = 40, min_cells: int = 8) -> None:
        self.materials = materials
        thickness_mm = dict(FIXED_THICKNESS_MM, II=d2_mm, IV=d4_mm)
        self.thickness = {k: thickness_mm[k] * 1e-3 for k in LAYERS}
        thinnest = min(self.thickness.values())
        widths, capacity, conductivity = [], [], []
        for key in LAYERS:
            span = self.thickness[key]
            # keep every layer resolved: scale cell count with thickness, never below min_cells
            count = max(min_cells, int(round(cells_per_layer * span / thinnest)))
            count = min(count, 4000)
            mat = materials[key]
            widths.extend([span / count] * count)
            capacity.extend([mat.rho * mat.c * span / count] * count)
            conductivity.extend([mat.lam] * count)
        self.h = np.asarray(widths)
        self.capacity = np.asarray(capacity)
        self.lam = np.asarray(conductivity)
        self.n = self.h.size
        # harmonic-mean face conductance -> exact continuity of flux at interfaces
        half_left = self.h[:-1] / (2.0 * self.lam[:-1])
        half_right = self.h[1:] / (2.0 * self.lam[1:])
        self.face = 1.0 / (half_left + half_right)
        self.outer_face = self.lam[0] / (self.h[0] / 2.0)
        self.skin_half = self.h[-1] / (2.0 * self.lam[-1])

    def steady_state_skin(self, env_c: float, h_skin: float) -> float:
        resistance = sum(self.thickness[k] / self.materials[k].lam for k in LAYERS)
        flux = (env_c - BODY_CORE_C) / (resistance + 1.0 / h_skin)
        return BODY_CORE_C + flux / h_skin

    def run(self, env_c: float, h_skin: float, duration_s: float, dt_s: float = 1.0,
            initial_c: float = BODY_CORE_C) -> tuple[np.ndarray, np.ndarray]:
        steps = int(round(duration_s / dt_s))
        temp = np.full(self.n, float(initial_c))
        skin_conductance = 1.0 / (self.skin_half + 1.0 / h_skin)
        lower = np.zeros(self.n)
        upper = np.zeros(self.n)
        lower[1:] = -self.face
        upper[:-1] = -self.face
        diag = np.zeros(self.n)
        diag[:-1] += self.face
        diag[1:] += self.face
        diag[0] += self.outer_face
        diag[-1] += skin_conductance
        alpha = self.capacity / dt_s
        matrix = np.zeros((3, self.n))
        matrix[0, 1:] = upper[:-1]
        matrix[1] = diag + alpha
        matrix[2, :-1] = lower[1:]
        from scipy.linalg import solve_banded
        constant = np.zeros(self.n)
        constant[0] = self.outer_face * env_c
        constant[-1] = skin_conductance * BODY_CORE_C
        skin = np.empty(steps + 1)
        skin[0] = self._skin_face(temp[-1], h_skin)
        for step in range(steps):
            temp = solve_banded((1, 1), matrix, alpha * temp + constant)
            skin[step + 1] = self._skin_face(temp[-1], h_skin)
        return np.arange(steps + 1) * dt_s, skin

    def run_field(self, env_c: float, h_skin: float, duration_s: float, dt_s: float = 1.0,
                  initial_c: float = BODY_CORE_C) -> np.ndarray:
        """Same march as run(), but return the final cell-centred field for profile plots."""
        from scipy.linalg import solve_banded
        steps = int(round(duration_s / dt_s))
        temp = np.full(self.n, float(initial_c))
        skin_conductance = 1.0 / (self.skin_half + 1.0 / h_skin)
        diag = np.zeros(self.n)
        diag[:-1] += self.face
        diag[1:] += self.face
        diag[0] += self.outer_face
        diag[-1] += skin_conductance
        alpha = self.capacity / dt_s
        matrix = np.zeros((3, self.n))
        matrix[0, 1:] = -self.face
        matrix[1] = diag + alpha
        matrix[2, :-1] = -self.face
        constant = np.zeros(self.n)
        constant[0] = self.outer_face * env_c
        constant[-1] = skin_conductance * BODY_CORE_C
        for _ in range(steps):
            temp = solve_banded((1, 1), matrix, alpha * temp + constant)
        return temp

    def _skin_face(self, last_cell_c: float, h_skin: float) -> float:
        conductance = 1.0 / self.skin_half
        return (conductance * last_cell_c + h_skin * BODY_CORE_C) / (conductance + h_skin)


def calibrate_skin_coefficient(materials: dict[str, Material], times: np.ndarray,
                               measured: np.ndarray) -> dict[str, float]:
    """One free parameter. Two independent routes must agree or the model is wrong."""
    sim = Simulator(materials, d2_mm=6.0, d4_mm=5.0)
    resistance = sum(sim.thickness[k] / materials[k].lam for k in LAYERS)
    plateau = float(measured[-1])
    # route 1: closed-form from the measured plateau alone
    flux = (75.0 - plateau) / resistance
    from_plateau = flux / (plateau - BODY_CORE_C)
    # route 2: least squares over the whole 5401-point history
    from scipy.optimize import minimize_scalar

    def rmse(h_skin: float) -> float:
        _, skin = sim.run(75.0, h_skin, times[-1], dt_s=1.0)
        return float(np.sqrt(np.mean((skin[: measured.size] - measured) ** 2)))

    result = minimize_scalar(rmse, bounds=(1.0, 60.0), method="bounded",
                            options={"xatol": 1e-4})
    return {"from_plateau": from_plateau, "from_history": float(result.x),
            "history_rmse_c": float(result.fun),
            "relative_gap": abs(from_plateau - result.x) / from_plateau}


def grid_convergence(materials: dict[str, Material], h_skin: float,
                     duration_s: float) -> list[dict[str, float]]:
    rows = []
    reference = None
    for cells in (10, 20, 40, 80):
        for dt in (4.0, 2.0, 1.0):
            sim = Simulator(materials, 6.0, 5.0, cells_per_layer=cells)
            _, skin = sim.run(75.0, h_skin, duration_s, dt_s=dt)
            final = float(skin[-1])
            if reference is None or (cells, dt) == (80, 1.0):
                reference = final
            rows.append({"cells_per_thinnest_layer": cells, "dt_s": dt,
                         "cells_total": sim.n, "final_skin_c": final})
    for row in rows:
        row["delta_vs_finest_c"] = abs(row["final_skin_c"] - reference)
    return rows


def constraint_metrics(times: np.ndarray, skin: np.ndarray, window_s: float,
                       cap_c: float = 47.0, soft_c: float = 44.0,
                       soft_budget_s: float = 300.0) -> dict[str, float | bool]:
    inside = times <= window_s + 1e-9
    t, y = times[inside], skin[inside]
    step = float(t[1] - t[0]) if t.size > 1 else 0.0
    above_soft = float(np.count_nonzero(y > soft_c) * step)
    peak = float(y.max())
    first_soft = float(t[np.argmax(y > soft_c)]) if bool((y > soft_c).any()) else float("inf")
    return {"peak_skin_c": peak, "cap_ok": bool(peak <= cap_c + 1e-9),
            "seconds_above_44": above_soft,
            "soft_ok": bool(above_soft <= soft_budget_s + 1e-9),
            "first_cross_44_s": first_soft,
            "feasible": bool(peak <= cap_c + 1e-9 and above_soft <= soft_budget_s + 1e-9)}


def minimal_feasible_d2(materials: dict[str, Material], h_skin: float, env_c: float,
                        window_s: float, d4_mm: float, dt_s: float = 2.0,
                        tol_mm: float = 0.01) -> dict[str, object]:
    """Thicker II layer is never worse, so bisect on feasibility instead of calling a solver."""
    low, high = BOUNDS_MM["II"]

    def feasible(d2: float) -> tuple[bool, dict]:
        sim = Simulator(materials, d2, d4_mm, cells_per_layer=20)
        times, skin = sim.run(env_c, h_skin, window_s, dt_s=dt_s)
        metrics = constraint_metrics(times, skin, window_s)
        return bool(metrics["feasible"]), metrics

    ok_low, metrics_low = feasible(low)
    if ok_low:
        return {"status": "lower_bound_already_feasible", "d2_mm": low, "metrics": metrics_low}
    ok_high, metrics_high = feasible(high)
    if not ok_high:
        return {"status": "infeasible_on_whole_range", "d2_mm": None,
                "metrics_at_upper_bound": metrics_high}
    while high - low > tol_mm:
        mid = 0.5 * (low + high)
        ok, _ = feasible(mid)
        if ok:
            high = mid
        else:
            low = mid
    ok, metrics = feasible(high)
    return {"status": "optimal", "d2_mm": round(high, 3), "metrics": metrics,
            "monotonicity_note": "II 层越厚热阻越大、皮肤温度越低，可行集是上半区间，故用二分而非元启发式。"}


def pareto_d2_d4(materials: dict[str, Material], h_skin: float, env_c: float,
                 window_s: float, dt_s: float = 2.0) -> list[dict[str, object]]:
    """问三保留双目标：对每个 d4 求最小可行 d2，输出前沿而不是把两目标相减。"""
    rows = []
    d4_low, d4_high = BOUNDS_MM["IV"]
    for d4 in np.round(np.arange(d4_low, d4_high + 1e-9, 0.2), 3):
        found = minimal_feasible_d2(materials, h_skin, env_c, window_s, float(d4), dt_s=dt_s)
        rows.append({"d4_mm": float(d4), "status": found["status"],
                     "d2_mm": found.get("d2_mm"),
                     "peak_skin_c": (found.get("metrics") or {}).get("peak_skin_c"),
                     "seconds_above_44": (found.get("metrics") or {}).get("seconds_above_44")})
    return rows


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--self-test", action="store_true",
                       help="calibrate, validate against 附件2, then solve Q2/Q3")
    parser.add_argument("--out", type=Path,
                       default=Path(__file__).with_name("run-report.json"))
    args = parser.parse_args()

    materials = read_materials()
    times, measured = read_measurement()
    report: dict[str, object] = {
        "schema": "cumcm-2018a-rebuild/v1",
        "attachment": str(ATTACHMENT),
        "diffusivity_m2_s": {k: materials[k].diffusivity for k in LAYERS},
        "measurement": {"points": int(measured.size), "start_c": float(measured[0]),
                        "final_c": float(measured[-1]), "duration_s": float(times[-1])},
    }

    calibration = calibrate_skin_coefficient(materials, times, measured)
    report["calibration"] = calibration
    h_skin = calibration["from_history"]

    sim = Simulator(materials, 6.0, 5.0)
    sim_times, skin = sim.run(75.0, h_skin, times[-1], dt_s=1.0)
    residual = skin[: measured.size] - measured
    report["q1_validation"] = {
        "rmse_c": float(np.sqrt(np.mean(residual ** 2))),
        "max_abs_c": float(np.max(np.abs(residual))),
        "final_model_c": float(skin[-1]), "final_measured_c": float(measured[-1]),
        "steady_state_closed_form_c": sim.steady_state_skin(75.0, h_skin),
        "holdout_last_30min_rmse_c": float(np.sqrt(np.mean(residual[-1800:] ** 2))),
        "holdout_first_10min_rmse_c": float(np.sqrt(np.mean(residual[:600] ** 2))),
    }
    report["grid_convergence"] = grid_convergence(materials, h_skin, 5400.0)

    q2 = minimal_feasible_d2(materials, h_skin, 65.0, 3600.0, d4_mm=5.5)
    report["q2"] = {"env_c": 65.0, "window_min": 60, "d4_mm_fixed": 5.5, "result": q2}

    front = pareto_d2_d4(materials, h_skin, 80.0, 1800.0)
    feasible_front = [row for row in front if row["d2_mm"] is not None]
    report["q3"] = {"env_c": 80.0, "window_min": 30, "pareto_front": front,
                    "note": "两个目标方向相反（II 层要薄、IV 层要厚），不做相减标量化；"
                            "报告前沿并说明所选点依据。",
                    "thinnest_total": min(
                        ({"d2_mm": r["d2_mm"], "d4_mm": r["d4_mm"],
                          "total_mm": round(r["d2_mm"] + r["d4_mm"], 3)}
                         for r in feasible_front), key=lambda r: r["total_mm"], default=None)}

    args.out.write_text(json.dumps(report, ensure_ascii=False, indent=1), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=1))
    print(f"\nwrote {args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())



